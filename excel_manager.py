# loading, saving, crud

import pandas as pd
from openpyxl import load_workbook, Woorkbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import datetime as datetime 
import os

EXCEL_FILE = "data.xls"

COLUMNS = ["Процедура", "Пояснение", "Формат", "Наименование поля", "Номер поля", "Публичный синоним", "grant", "Статус разработки", "Дата обновления"]

STATUS_COLORS = {"Архив": "FFC7CE", "Готово к внедрению": "FFEB9C", "Используется": "C6EFCE"}

# Загрузка данных в pandas DataFrame"""
def load_data():
    try:
        return pd.read_excel(EXCEL_FILE)
    except FileNotFoundError:
        return pd.DataFrame(columns=COLUMNS)

# Красим строку в зависимости от статуса разработки
def _apply_formatting(ws, row_idx, status_value):
    if status_value in STATUS_COLORS:
        fill = PatternFill(start_color=STATUS_COLORS[status_value], end_color=STATUS_COLORS[status_value], fill_type="solid")
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

# Добавляем запись в таблицу с сохранением форматирования
def add_record(record: dict):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except FileNotFoundError:
        # если файла ещё нет, создаём новый через pandas
        df = pd.DataFrame(columns=COLUMNS)
        df.to_excel(EXCEL_FILE, index=False)
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    new_row_idx = ws.max_row + 1

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=new_row_idx, column=col_idx).value = record.get(col_name, "")

    # применяем форматирование
    _apply_formatting(ws, new_row_idx, record.get("Статус разработки", ""))

    wb.save(EXCEL_FILE)

# Обновить запись (row_index от 0, как в pandas)
def update_record(row_index: int, record: dict):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    excel_row = row_index + 2 # +2: 1 строка заголовков + сдвиг с 0

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=excel_row, column=col_idx).value = record.get(col_name, "")

    # обновляем цвет
    _apply_formatting(ws, excel_row, record.get("Статус разработки", ""))

    wb.save(EXCEL_FILE)

# Удалить запись по индексу (row_index от 0, как в pandas)
def delete_record(row_index: int):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    excel_row = row_index + 2 # +2: 1 строка заголовков
    ws.delete_rows(excel_row, 1)

    wb.save(EXCEL_FILE)
